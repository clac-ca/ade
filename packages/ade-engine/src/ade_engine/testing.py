"""Small test helpers for ADE engine outputs."""

from pathlib import Path

from openpyxl import load_workbook
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet


def open_workbook(output_path: Path) -> Workbook:
    return load_workbook(output_path)


def _select_sheet(workbook: Workbook, sheet: str | int | None) -> Worksheet:
    if sheet is None:
        return workbook.active
    if isinstance(sheet, int):
        return workbook.worksheets[sheet]
    return workbook[sheet]


def read_rows(
    output_path: Path, *, sheet: str | int | None = None
) -> list[list[object]]:
    workbook = load_workbook(output_path, data_only=True)
    worksheet = _select_sheet(workbook, sheet)

    rows: list[list[object]] = []
    for values in worksheet.iter_rows(values_only=True):
        row = list(values)
        while row and row[-1] is None:
            row.pop()
        if not row:
            continue
        rows.append(row)

    return rows
