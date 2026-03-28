"""Runtime entrypoint for the ADE engine."""

from collections.abc import Iterable
import csv
from dataclasses import dataclass, field
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from ade_engine.config import EngineConfig


@dataclass(frozen=True)
class ValidationIssue:
    row_index: int
    field: str
    message: str


@dataclass(frozen=True)
class RunResult:
    output_path: Path
    validation_issues: list[ValidationIssue] = field(default_factory=list)


@dataclass(frozen=True)
class _Row:
    values: list[object]
    source_index: int


@dataclass(frozen=True)
class _Column:
    header: str | None
    sample_values: list[object]


@dataclass
class _HookContext:
    workbook: Workbook
    worksheet: Worksheet


def _read_csv_rows(input_path: Path) -> list[list[object]]:
    with input_path.open(newline="", encoding="utf-8") as file:
        return [list(row) for row in csv.reader(file)]


def _read_sheet_rows(worksheet: Worksheet) -> list[list[object]]:
    return [list(row) for row in worksheet.iter_rows(values_only=True)]


def _load_input_rows(input_path: Path) -> list[list[object]]:
    suffix = input_path.suffix.lower()

    if suffix == ".csv":
        return _read_csv_rows(input_path)

    if suffix in {".xlsx", ".xlsm"}:
        workbook = load_workbook(input_path, data_only=True)
        return _read_sheet_rows(workbook.active)

    raise ValueError(
        f"Unsupported input file type '{input_path.suffix}' for '{input_path.name}'.",
    )


def _trimmed_row(values: Iterable[object]) -> list[object]:
    row = list(values)
    while row and row[-1] in (None, ""):
        row.pop()
    return row


def _row_is_empty(values: list[object]) -> bool:
    return all(value in (None, "") for value in values)


def _score_row(detectors: list, values: list[object]) -> float:
    row = _Row(values=values, source_index=0)
    return max((float(detector(row)) for detector in detectors), default=0.0)


def _detect_header_index(config: EngineConfig, rows: list[list[object]]) -> int:
    best_index = -1
    best_score = 0.0

    for index, values in enumerate(rows):
        score = _score_row(config.row_detectors["header"], values)
        if score > best_score:
            best_index = index
            best_score = score

    if best_index < 0:
        raise ValueError("Unable to detect a header row.")

    return best_index


def _collect_data_rows(
    config: EngineConfig,
    rows: list[list[object]],
    *,
    header_index: int,
) -> list[_Row]:
    detected_rows: list[_Row] = []
    in_table = False

    for source_index in range(header_index + 1, len(rows)):
        values = rows[source_index]

        if _row_is_empty(values):
            if in_table:
                break
            continue

        score = _score_row(config.row_detectors["data"], values)
        if score <= 0:
            if in_table:
                break
            continue

        in_table = True
        detected_rows.append(_Row(values=values, source_index=source_index + 1))

    return detected_rows


def _column_value(row: _Row, index: int) -> object:
    return row.values[index] if index < len(row.values) else None


def _build_columns(headers: list[object], data_rows: list[_Row]) -> list[_Column]:
    columns: list[_Column] = []
    column_count = max(
        len(headers), max((len(row.values) for row in data_rows), default=0)
    )

    for index in range(column_count):
        header = headers[index] if index < len(headers) else None
        sample_values = [_column_value(row, index) for row in data_rows]
        columns.append(
            _Column(
                header=str(header) if header not in (None, "") else None,
                sample_values=sample_values,
            )
        )

    return columns


def _map_fields(config: EngineConfig, columns: list[_Column]) -> dict[str, int]:
    field_order = {field_name: index for index, field_name in enumerate(config.fields)}
    candidates: list[tuple[float, int, int, str]] = []

    for field_name, rules in config.fields.items():
        for column_index, column in enumerate(columns):
            score = sum(float(detector(column)) for detector in rules.detectors)
            if score <= 0:
                continue
            candidates.append((score, field_order[field_name], column_index, field_name))

    assignments: dict[str, int] = {}
    claimed_columns: set[int] = set()
    claimed_fields: set[str] = set()

    for _, _, column_index, field_name in sorted(
        candidates,
        key=lambda candidate: (-candidate[0], candidate[1], candidate[2]),
    ):
        if field_name in claimed_fields or column_index in claimed_columns:
            continue
        assignments[field_name] = column_index
        claimed_fields.add(field_name)
        claimed_columns.add(column_index)

    return assignments


def _normalize_value(config: EngineConfig, field_name: str, value: object) -> object:
    normalized = value
    for transform in config.fields[field_name].transforms:
        normalized = transform(normalized)
    return normalized


def _validate_rows(
    config: EngineConfig,
    normalized_rows: list[dict[str, object]],
    source_rows: list[_Row],
) -> list[ValidationIssue]:
    issues: list[ValidationIssue] = []

    for row, source_row in zip(normalized_rows, source_rows, strict=True):
        for field_name, value in row.items():
            for validator in config.fields[field_name].validators:
                message = validator(value)
                if message is None:
                    continue
                issues.append(
                    ValidationIssue(
                        row_index=source_row.source_index,
                        field=field_name,
                        message=str(message),
                    )
                )

    return issues


def _write_output_workbook(
    config: EngineConfig,
    normalized_rows: list[dict[str, object]],
) -> Workbook:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Output"

    field_names = list(config.fields)
    worksheet.append(field_names)

    for row in normalized_rows:
        worksheet.append([row.get(field_name) for field_name in field_names])

    context = _HookContext(
        workbook=workbook,
        worksheet=worksheet,
    )
    for hook in config.hooks.get("on_table_written", []):
        hook(context)

    return workbook


def process(*, config: EngineConfig, input_path: Path, output_dir: Path) -> RunResult:
    if not input_path.exists():
        raise FileNotFoundError(f"Input path does not exist: '{input_path}'.")

    if input_path.is_dir():
        raise ValueError("Directory input is not supported yet.")

    if not input_path.is_file():
        raise ValueError(f"Input path must be a regular file: '{input_path}'.")

    output_dir.mkdir(parents=True, exist_ok=True)

    rows = [_trimmed_row(row) for row in _load_input_rows(input_path)]
    header_index = _detect_header_index(config, rows)
    data_rows = _collect_data_rows(config, rows, header_index=header_index)
    columns = _build_columns(rows[header_index], data_rows)
    field_map = _map_fields(config, columns)

    normalized_rows: list[dict[str, object]] = []
    for source_row in data_rows:
        normalized_row: dict[str, object] = {}
        for field_name in config.fields:
            column_index = field_map.get(field_name)
            raw_value = (
                _column_value(source_row, column_index)
                if column_index is not None
                else None
            )
            normalized_row[field_name] = _normalize_value(config, field_name, raw_value)
        normalized_rows.append(normalized_row)

    validation_issues = _validate_rows(config, normalized_rows, data_rows)
    workbook = _write_output_workbook(config, normalized_rows)

    output_path = output_dir / f"{input_path.stem}.normalized.xlsx"
    workbook.save(output_path)
    return RunResult(output_path=output_path, validation_issues=validation_issues)
